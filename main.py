import re
import os
import sys
from window_ui import Ui_MainWindow
from openpyxl import load_workbook
from PyQt6 import QtWidgets
from PyQt6.QtWidgets import QTableWidgetItem, QMessageBox
from PyQt6.QtCore import Qt


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.folder = None
        self.barcode_column = self.barcode_col_box.value()
        self.warehouse = {}
        self.sorted_warehouse = {}

        self.unsorted_table.setColumnWidth(0, 50)

        self.open_btn.triggered.connect(self.open_file)

    def open_file(self):
        self.folder = QtWidgets.QFileDialog.getExistingDirectory()

        if self.folder:
            self.parsing_data()
            self.count_reports()

    def count_reports(self):
        unsorted_table = self.unsorted_table
        sorted_table = self.sorted_table
        unsorted_table.setRowCount(len(self.warehouse))
        i = 0

        for alley in self.warehouse:
            alley_item = QTableWidgetItem(alley)
            rows_item = QTableWidgetItem(str(self.warehouse[alley]['rows']))
            pilot_item = QTableWidgetItem(self.warehouse[alley]['pilot'])
            unread_item = QTableWidgetItem(str(self.warehouse[alley]['unread']) + '%')

            alley_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            rows_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            pilot_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            unread_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            unsorted_table.setItem(i, 0, alley_item)
            unsorted_table.setItem(i, 1, rows_item)
            unsorted_table.setItem(i, 2, pilot_item)
            unsorted_table.setItem(i, 3, unread_item)
            i += 1

            if self.warehouse[alley]['pilot'] not in self.sorted_warehouse:
                self.sorted_warehouse[self.warehouse[alley]['pilot']] = \
                    {'rows': self.warehouse[alley]['rows'],
                     'unread': self.warehouse[alley]['unread']}
            elif self.warehouse[alley]['pilot'] in self.sorted_warehouse:
                current_pilot = self.sorted_warehouse[self.warehouse[alley]['pilot']]
                current_unread = int(current_pilot['rows']) * float(current_pilot['unread'])
                new_unread = int(self.warehouse[alley]['rows']) * float(self.warehouse[alley]['unread'])
                all_rows = int(current_pilot['rows']) + int(self.warehouse[alley]['rows'])
                current_pilot['rows'] = int(current_pilot['rows']) + int(self.warehouse[alley]['rows'])
                current_pilot['unread'] = round((current_unread + new_unread) / all_rows, 2)

        sorted_table.setRowCount(len(self.sorted_warehouse))
        i = 0
        for pilot in self.sorted_warehouse:
            rows_item = QTableWidgetItem(str(self.sorted_warehouse[pilot]['rows']))
            unread_item = QTableWidgetItem(str(self.sorted_warehouse[pilot]['unread']) + '%')
            pilot_item = QTableWidgetItem(pilot)

            rows_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            unread_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            pilot_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            sorted_table.setItem(i, 0, pilot_item)
            sorted_table.setItem(i, 1, rows_item)
            sorted_table.setItem(i, 2, unread_item)

            i += 1

    def parsing_data(self):
        self.barcode_column = self.barcode_col_box.value()
        self.unsorted_table.setRowCount(0)
        self.sorted_table.setRowCount(0)
        self.warehouse = {}
        self.sorted_warehouse = {}

        row_count = 0
        for file in os.listdir(self.folder):
            unread_count = 0
            if re.match(r'.+_[а-яА-Я]+_(\d{1,2}\.){2}\d{4}', file) is not None:
                sheet = load_workbook(self.folder + '\\' + file)["Выгрузка провайдера"]
                name = file[:-5]
                alley_name = re.split(r'_[а-яА-Я]+_', name)[0]
                pilot_name = re.search(r'[а-яА-Я]+', name)[0]
                date = re.split(r'_[а-яА-Я]+_', name)[1]
                # Поиск количества паллет
                if sheet['A' + str(len(sheet['A']))].value:
                    if str(sheet.cell(column=1, row=len(sheet['A'])).value).isdigit():
                        row_count = int(sheet['A' + str(len(sheet['A']))].value)
                    else:
                        row_count = 0
                else:
                    if str(sheet.cell(column=1, row=len(sheet['A']) - 1).value).isdigit():
                        row_count = int(sheet['A' + str(len(sheet['A']) - 1)].value)
                for i in range(2, row_count + 1):
                    if sheet.cell(row=i, column=self.barcode_column).value == "UNREADABLE":
                        unread_count += 1
                if unread_count != 0:
                    unread_percent = round(unread_count / row_count * 100, 2)
                else:
                    unread_percent = 0
                self.warehouse[alley_name] = {'pilot': pilot_name,
                                              'date': date,
                                              'rows': row_count,
                                              'unread': unread_percent}


def main():
    sys.excepthook = my_excepthook
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec()


def my_excepthook(type_error, value, t_back):
    QMessageBox.critical(MainWindow(), "CRITICAL ERROR", str(value))
    sys.__excepthook__(type_error, value, t_back)


if __name__ == '__main__':
    main()
